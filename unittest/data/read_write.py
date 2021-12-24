"""
@File:read_write.py
@DateTime:2021/12/18 20:17
@Author:Ben
@Desc:
"""
import openpyxl


class ReadWrite:
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet
        self.wb = openpyxl.load_workbook(self.file)
        self.table = self.wb[self.sheet]
        self.table = self.wb.active
        self.max_row = self.table.max_row
        self.max_col = self.table.max_column
        pass

    def read(self):
        list2 = []
        for row in range(2, self.max_row + 1):
            list1 = []
            for col in range(1, self.max_col + 1):
                content = self.table.cell(row, col).value
                list1.append(content)
            list2.append(list1)
        self.wb.close()
        return list2

    def write(self, *arg):
        for col in range(1, len(arg) + 1):
            self.table.cell(self.max_row + 1, col).value = arg[col - 1]  # 将arg元祖的元素按照索引依次写入Excel表中
        self.wb.save(self.file)
        self.wb.close()


if __name__ == '__main__':
    file = r"D:\PyCharm\Python_Automation\web_automation\day_07\data\testdata.xlsx"
    sheet = "login"
    doc1 = ReadWrite(file, sheet)
    print(doc1.read())
    doc1.write("shelly", "123456")
